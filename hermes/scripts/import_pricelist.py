"""Import a distributor Excel/CSV price list into the Odoo catalog.

Per brand (configured under pricebook.brands): upsert products keyed on the
manufacturer part number stored as the product's INTERNAL REFERENCE
(default_code); cost -> product.supplierinfo under the brand's vendor partner;
sale price = cost * (1 + markup_pct/100) -> list_price.

DEFAULTS TO DRY-RUN: prints the would-be creates/updates, writes nothing.

Usage:  python scripts/import_pricelist.py --brand acme path/to/list.xlsx [--live] [--odoo-db NAME]
"""
from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from core.config import load_config           # noqa: E402
from core.actions import _now                 # noqa: E402
from core.product_matcher import norm_code    # noqa: E402
from connectors.odoo_client import OdooClient, OdooError        # noqa: E402
from connectors.sheets_client import SheetsClient, SheetsError  # noqa: E402


def _num(v) -> float | None:
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None


def read_pricelist(xlsx_bytes: bytes, brand: dict) -> list[dict]:
    """Map the brand's columns; skip rows without part or cost; compute sale price."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr_i = brand.get("header_row", 1) - 1
    header = [str(c or "").strip().lower() for c in rows[hdr_i]]
    want = {k: str(v).strip().lower() for k, v in brand["columns"].items()}
    idx = {}
    for key, label in want.items():
        if label not in header:
            raise SystemExit(f"Column '{brand['columns'][key]}' not found in header row {hdr_i + 1}: {header}")
        idx[key] = header.index(label)

    markup = 1 + (brand.get("markup_pct", 25) / 100.0)
    out = []
    for r in rows[hdr_i + 1:]:
        part = str(r[idx["part"]] or "").strip() if len(r) > idx["part"] else ""
        cost = _num(r[idx["cost"]]) if len(r) > idx["cost"] else None
        if not part or cost is None:
            continue
        desc = str(r[idx["description"]] or "").strip() if len(r) > idx["description"] else ""
        out.append({"part": part, "description": desc, "cost": cost,
                    "sale_price": round(cost * markup, 2)})
    return out


def upsert(odoo: OdooClient, sheets, audit_tab, vendor_id, rows, dry) -> tuple[int, int]:
    run_mode = "dry-run" if dry else "live"
    existing = odoo.search_read("product.product", [], ["default_code"], limit=100000)
    by_code = {norm_code(p.get("default_code")): p["id"] for p in existing if p.get("default_code")}

    created = updated = 0
    for row in rows:
        key = norm_code(row["part"])
        pid = by_code.get(key)
        action = "update" if pid else "create"
        if dry:
            print(f"  [SIM] {action}: {row['part']:<20} cost={row['cost']:<10} sale={row['sale_price']}")
        else:
            if pid is None:
                pid = odoo.create_product(row["description"] or row["part"],
                                          default_code=row["part"], list_price=row["sale_price"])
                created += 1
            else:
                odoo.execute("product.product", "write", [pid], {"list_price": row["sale_price"]})
                updated += 1
            odoo.upsert_supplierinfo(odoo.product_tmpl_id(pid), vendor_id, row["cost"])
    sheets.append_row(audit_tab, [_now(), "pricebook", "import_pricelist",
                                  f"{len(rows)} row(s): {created} created, {updated} updated",
                                  "dry-run" if dry else "ok", run_mode])
    return created, updated


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--brand", required=True)
    ap.add_argument("--live", action="store_true")
    ap.add_argument("--odoo-db", default=None)
    args = ap.parse_args()

    cfg = load_config()
    if args.odoo_db:
        cfg["odoo"]["db"] = args.odoo_db
    dry = cfg.get("runtime", {}).get("dry_run", True) and not args.live

    brands = cfg.get("pricebook", {}).get("brands", {})
    if args.brand not in brands:
        print(f"Unknown brand '{args.brand}'. Configured: {', '.join(brands) or '(none)'}")
        return 1
    brand = brands[args.brand]
    rows = read_pricelist(Path(args.file).read_bytes(), brand)
    print(f"Price list '{args.brand}': {len(rows)} usable row(s)   "
          f"markup {brand.get('markup_pct', cfg['pricebook'].get('default_markup_pct', 25))}%   "
          f"mode: {'DRY-RUN' if dry else 'LIVE'}")

    try:
        odoo = OdooClient.from_config(cfg)
        sheets = SheetsClient.from_config(cfg)
    except (OdooError, SheetsError) as exc:
        print(f"FAILED (connect): {exc}")
        return 1

    vendor_id = 0
    if not dry:
        vendor_id = odoo.ensure_vendor(brand["vendor_name"])
    created, updated = upsert(odoo, sheets, cfg["sheets"]["tabs"]["audit"], vendor_id, rows, dry)
    if not dry:
        print(f"  -> {created} created, {updated} updated (vendor '{brand['vendor_name']}')")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
