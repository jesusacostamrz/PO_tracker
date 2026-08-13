"""Local distributor pricebooks: big xlsx price lists the RFQ pipeline consults
directly instead of importing 70k+ products into Odoo (a full import would be
hours of API calls for parts we may never quote). A queued RFQ line that hits a
pricebook is priced at the LIST price — no extra margin unless the email
instructs one; an email-granted discount multiplies (10% -> x0.9), capped by
pricebook.max_discount_pct — and its product is auto-created on demand by
quote_actions (name = the distributor TYPE designation, item number kept in the
sale description).

Config (pricebook.files.<key>): path (relative to hermes/), vendor_name,
currency, mxn_fx_surcharge, header_row, columns {item, type, list, cost}.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from core.product_matcher import norm_code

_ROOT = Path(__file__).resolve().parents[1]


def _clean_type(s) -> str:
    """Distributor lists have sloppy spacing ('NS 35/ 7,5', 'FBSK  3-7,5') —
    normalize to the official designation for product names."""
    s = re.sub(r"\s*/\s*", "/", str(s or ""))
    return re.sub(r"\s{2,}", " ", s).strip()


def _num(v) -> float | None:
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None


def _colsig(fcfg: dict) -> str:
    return json.dumps({"header_row": fcfg.get("header_row", 1),
                       "columns": fcfg["columns"]}, sort_keys=True)


def _read_rows(path: Path, fcfg: dict) -> list[list]:
    """[[item, type, list, cost], ...] — cached as JSON next to the xlsx
    (70k rows via openpyxl is tens of seconds; the cache loads in <1s). The
    cache is invalidated by a newer xlsx OR a changed column mapping — a
    repointed columns.list must never keep serving old prices."""
    cache = path.with_suffix(path.suffix + ".cache.json")
    if cache.exists() and cache.stat().st_mtime >= path.stat().st_mtime:
        try:
            data = json.loads(cache.read_text(encoding="utf-8"))
            if isinstance(data, dict) and data.get("colsig") == _colsig(fcfg):
                return data["rows"]
        except (json.JSONDecodeError, KeyError):
            pass  # corrupt/old-format cache -> rebuild from the xlsx

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    hdr_i = fcfg.get("header_row", 1)
    want = {k: str(v).strip().lower() for k, v in fcfg["columns"].items()}
    idx: dict[str, int] = {}
    rows: list[list] = []
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < hdr_i:
            continue
        if i == hdr_i:
            header = [str(c or "").strip().lower() for c in r]
            for key, label in want.items():
                if label not in header:
                    raise ValueError(f"pricebook {path.name}: column '{fcfg['columns'][key]}' "
                                     f"not in header row {hdr_i}: {header}")
                idx[key] = header.index(label)
            continue
        cell = lambda k: r[idx[k]] if len(r) > idx[k] else None  # noqa: E731
        item = str(cell("item") or "").strip()
        type_ = str(cell("type") or "").strip()
        lst = _num(cell("list"))
        if not (item or type_) or lst is None or lst <= 0:
            continue
        rows.append([item, type_, lst, _num(cell("cost"))])
    wb.close()
    cache.write_text(json.dumps({"colsig": _colsig(fcfg), "rows": rows}), encoding="utf-8")
    return rows


class Pricebook:
    def __init__(self, key: str, fcfg: dict, rows: list[list]):
        self.key = key
        self.vendor_name = fcfg.get("vendor_name") or key
        self.currency = str(fcfg.get("currency") or "USD").upper()
        self.mxn_fx_surcharge = float(fcfg.get("mxn_fx_surcharge") or 0)
        self.image_site = str(fcfg.get("image_site") or "").strip() or None
        self.index: dict[str, dict] = {}
        for item, type_, lst, cost in rows:
            rec = {"item": item, "type": _clean_type(type_), "list": lst, "cost": cost}
            for k in (norm_code(item), norm_code(type_)):
                if k:
                    self.index.setdefault(k, rec)

    def lookup(self, part_number, description="") -> dict | None:
        """Exact normalized match on the RFQ part# (against item OR type), else an
        unambiguous type/item inside the description. Multi-word type designations
        ("FBS 10-5 BU") only exist as JOINED adjacent-token windows — longest
        window wins, so a full type beats its own prefix ("FBS 10-5" is a
        DIFFERENT product). This also rescues lines whose cited item# is a typo
        but whose description carries the real type."""
        code = norm_code(part_number)
        if code and code in self.index:
            return self.index[code]
        # no comma in the split — PXC types embed commas ("DFK-2,8"); edge
        # punctuation is stripped per token instead
        toks = [t for t in (tok.strip(",.;:")
                            for tok in re.split(r"[\s;:()\[\]\"']+", str(description or "")))
                if norm_code(t)]
        for size in (4, 3, 2, 1):
            hits: dict[str, dict] = {}
            for i in range(len(toks) - size + 1):
                t = norm_code("".join(toks[i:i + size]))
                if len(t) >= 5 and any(c.isdigit() for c in t) and t in self.index:
                    hits[self.index[t]["item"] or t] = self.index[t]
            if len(hits) == 1:
                return next(iter(hits.values()))
            if len(hits) > 1:
                return None  # two different catalog parts at the same length -> ambiguous
        return None


_memo: dict[tuple, Pricebook] = {}  # (key, path, mtime, colsig) -> built book


def load_pricebooks(cfg: dict) -> list[Pricebook]:
    out: list[Pricebook] = []
    for key, fcfg in (cfg.get("pricebook", {}).get("files") or {}).items():
        path = Path(fcfg["path"])
        if not path.is_absolute():
            path = _ROOT / path
        if not path.exists():
            print(f"[pricebook] {key}: {path} not found — skipped")
            continue
        # memoized per process: intake batches / --watch loops call this per
        # message, and rebuilding a 72k-entry index each time is pure waste
        mk = (key, str(path), path.stat().st_mtime, _colsig(fcfg))
        if mk not in _memo:
            _memo[mk] = Pricebook(key, fcfg, _read_rows(path, fcfg))
        out.append(_memo[mk])
    return out


def apply_pricebook(matches, products: list[dict], cfg: dict) -> int:
    """Post-match_lines pass: resolve queued lines against the local pricebooks.

    A hit stashes the record on the line (``_pricebook``) for quote_actions to
    price (list x fx, email discount/margin). If the part already exists in the
    Odoo catalog under its type OR item number (e.g. the RFQ cited the item#
    but the product was created under the type), the existing product is reused
    — never a duplicate. Returns the number of lines resolved.
    """
    books = load_pricebooks(cfg)
    if not books:
        return 0
    by_code = {norm_code(p.get("name")): p for p in products if p.get("name")}
    by_code.update({norm_code(p.get("default_code")): p for p in products if p.get("default_code")})
    hits = 0
    for m in matches:
        if m.status != "queue":
            continue
        for pb in books:
            rec = pb.lookup(m.line.get("part_number"), m.line.get("description"))
            if not rec:
                continue
            m.line["_pricebook"] = {**rec, "currency": pb.currency,
                                    "mxn_fx_surcharge": pb.mxn_fx_surcharge,
                                    "vendor": pb.vendor_name, "image_site": pb.image_site}
            if not m.line.get("manufacturer"):
                m.line["manufacturer"] = pb.vendor_name  # image-search hint
            prod = by_code.get(norm_code(rec["item"])) or by_code.get(norm_code(rec["type"]))
            if prod:
                m.status, m.product, m.score = "matched", prod, 100.0
                m.reason = f"pricebook {pb.key}: existing product"
            else:
                m.reason = f"pricebook {pb.key}: list price"
            hits += 1
            break
    return hits
