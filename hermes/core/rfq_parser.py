"""Parse a customer RFQ (component list) into structured line items via the LLM.

Inputs arrive in mixed formats — Excel/CSV/PDF attachments, images/photos, or a
table pasted in the email body. Everything is normalized to text (or image data-URLs)
and pushed through ONE extraction prompt, so all formats yield the same JSON.
"""
from __future__ import annotations

import base64
import io
import re

_MAX_TEXT_CHARS = 15000


def xlsx_to_text(data: bytes) -> str:
    """All sheets -> TSV-ish text. Header detection is the LLM's job, not ours."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"=== SHEET {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(c.strip() for c in cells):
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines).strip()


def _system_prompt(company: dict) -> str:
    aka = ", ".join(company.get("aka", []) or [])
    who = f'Our company is the SUPPLIER named "{company.get("legal_name", "")}"'
    if aka:
        who += f" (aka {aka})"
    return f"""You extract a Request For Quotation (RFQ) — a list of components a customer wants priced.

{who}. The RFQ is sent BY a customer TO us; the customer is the requesting company, never our own.

Return ONLY a JSON object (null where absent):
{{
  "customer_name": string,   // the requesting company, if identifiable; else null
  "rfq_ref": string,         // the customer's RFQ/requisition number, if any; else null
  "margin_pct": number,      // profit margin % EXPLICITLY instructed in the email body (e.g. "add 40%"); else null
  "price_source_site": string, // web domain the body EXPLICITLY orders us to take PRICES from ("include the price from <site>", "con precios de <sitio>"); a link given only as reference or for images is NOT a pricing instruction -> null
  "currency": string,        // ISO code (MXN, USD, ...) of the unit prices ON THE DOCUMENT — from an explicit currency label/legend; null if prices absent or currency not shown
  "line_items": [
    {{"part_number": string,  // manufacturer part number / catalog code; null if only a description
      "description": string,  // what the item is, as written
      "manufacturer": string, // brand/maker (e.g. NITRA, SMC, BIMBA) — from a brand column OR embedded in the description; else null
      "quantity": number,     // requested quantity; default 1 if truly absent
      "unit_cost": number,    // OUR unit cost — ONLY in supplier-quote mode (rule below); else null
      "cost_currency": string}} // ISO code for THIS line's unit_cost when shown per line (documents can mix USD and MXN lines); null if not shown or same as the top-level currency
  ]
}}

Rules:
- Extract EVERY requested item; never invent items; never drop rows.
- The EMAIL BODY may carry explicit instructions from OUR OWN salesperson. Those instructions
  OVERRIDE anything inferred from the attached document: "quote for <X>" / "cotización para <X>"
  names the customer (customer_name = X, resolve to the full company name if obvious); a stated
  profit margin ("increase 40%", "agrega 30% de margen") goes to margin_pct; an explicit order
  to take the prices from a website ("include the price of <site>") puts that site's domain in
  price_source_site.
- SUPPLIER-QUOTE MODE: applies whenever the attachment is a QUOTATION issued by ANOTHER company
  (their letterhead/folio, prices listed — a "Cotización", not a request) AND the body asks us to
  build a customer quote from it (it names a customer and/or a profit margin). The body does NOT
  need to say "supplier" — a margin instruction only makes sense against the document's prices.
  Then: customer_name comes ONLY from the body instruction — even a short or lowercase name
  ("quote for abc" -> customer_name "abc") — NEVER from the attached document: every company
  on it (the quotation's issuer, letterhead, addressee) is a supplier-side party, not the
  customer. If the body names no customer, customer_name is null. Each line's unit price on
  the document is OUR COST -> unit_cost (with its cost_currency when the document shows it
  per line), and rfq_ref is that quotation's folio/number.
- RFQs usually arrive FORWARDED by our own salespeople — ignore the forwarder. The customer is the
  ORIGINAL sender in the quoted forwarded header (its "From:"/"De:" line); take their company name,
  or infer it from their email domain (e.g. jperez@acme-corp.com -> Acme Corp).
- part_number is the manufacturer's MODEL/catalog code (e.g. "6204-2RS", "FRN15G15-4T", "FX3U-128M",
  "LC1D09"). It is often EMBEDDED mid-description ("VARIADOR FUJI ELECTRIC FRN15G15-4T 3 HP" ->
  part_number "FRN15G15-4T") — pull it out; the full text stays in description.
- NEVER use as part_number: line numbers, or regulatory/certification codes printed on labels
  (KCC-..., MSIP-..., UL, CE, NOM ...) — those are not catalog codes; leave them out entirely.
- Customer tables often ALSO carry the customer's own internal item/account/stock code (a column like
  "No. de parte ABC", "Código", "Item #", often sequential or same-prefixed for every row). That is NOT
  the part_number. Prefer the manufacturer/catalog code; if only the internal code exists, put it at the
  start of the description (e.g. "[cód. cliente 123456] ...") and set part_number null.
- Numbers: dot decimal, no thousands separators; quantity must be a number.
- Content may be Spanish/English or both. Ignore signatures, disclaimers, prices the customer
  typed (unit_cost in supplier-quote mode is the one exception).
- Output JSON only — no prose, no code fences."""


def _domain(v) -> str | None:
    """'https://www.Site.com/path' / 'Site.com' -> 'site.com'; junk -> None."""
    s = str(v or "").strip().lower()
    s = re.sub(r"^[a-z]+://", "", s).split("/")[0].split("?")[0]
    s = s.removeprefix("www.")
    return s if "." in s else None


def _img_data_url(img_bytes: bytes, filename: str) -> str:
    ext = filename.lower().rsplit(".", 1)[-1]
    mime = "image/png" if ext == "png" else "image/jpeg"
    return f"data:{mime};base64," + base64.b64encode(img_bytes).decode("ascii")


def parse_rfq(sources: list[tuple[str, str, bytes | str]], llm, company: dict) -> dict:
    """sources: [(kind, filename, payload)] with kind in {'xlsx','image','text','pdf'}.

    Text-ish sources are concatenated into one prompt. Any image routes the whole
    thing through the vision model, text included — the body often carries the
    customer identity/instructions while the image shows the items.
    """
    system = _system_prompt(company)
    texts: list[str] = []
    image_urls: list[str] = []
    for kind, filename, payload in sources:
        if kind == "xlsx":
            texts.append(f"--- ATTACHMENT {filename} ---\n{xlsx_to_text(payload)}")
        elif kind == "text":
            texts.append(f"--- EMAIL BODY ---\n{payload}")
        elif kind == "image":
            image_urls.append(_img_data_url(payload, filename))
        elif kind == "pdf":
            # reuse the PO pipeline's PDF machinery: text PDFs as text (cheap),
            # scanned PDFs rendered to images for the vision model
            import core.po_parser as po_parser
            txt = po_parser.extract_text(payload)
            if len(txt) >= po_parser._TEXT_MIN_CHARS:
                texts.append(f"--- ATTACHMENT {filename} ---\n{txt}")
            else:
                image_urls.extend(po_parser.render_pages_as_data_urls(payload))

    if image_urls:
        user_text = "Extract the RFQ from the attached image(s)."
        if texts:
            blob = "\n\n".join(texts)[:_MAX_TEXT_CHARS]
            user_text = ("Extract the RFQ from the text below AND the attached image(s) together. "
                         "The text may hold the customer identity, instructions, and/or items; "
                         "the image(s) may show the items themselves. "
                         "List each requested item once, even if it appears in both.\n\n" + blob)
        result = llm.vision_json(system=system, user_text=user_text,
                                 image_data_urls=image_urls[:4], max_tokens=16000)
        result["_source"] = "vision+text" if texts else "vision"
    elif texts:
        blob = "\n\n".join(texts)[:_MAX_TEXT_CHARS]
        result = llm.chat_json(system=system, user="RFQ content:\n\n" + blob, max_tokens=16000)
        result["_source"] = "text"
    else:
        result = {"customer_name": None, "rfq_ref": None, "line_items": [], "_source": "empty"}

    result["line_items"] = [li for li in (result.get("line_items") or []) if isinstance(li, dict)]

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    result["margin_pct"] = _f(result.get("margin_pct"))
    result["price_source_site"] = _domain(result.get("price_source_site"))
    cur = str(result.get("currency") or "").strip().upper()
    result["currency"] = cur if len(cur) == 3 and cur.isalpha() else None
    # normalize lines defensively — downstream indexes these keys
    for li in result["line_items"]:
        li["part_number"] = (li.get("part_number") or None)
        li["description"] = str(li.get("description") or "")
        li["manufacturer"] = str(li.get("manufacturer") or "")
        li["quantity"] = _f(li.get("quantity")) or 1.0
        li["unit_cost"] = _f(li.get("unit_cost"))
        lc = str(li.get("cost_currency") or "").strip().upper()
        li["cost_currency"] = lc if len(lc) == 3 and lc.isalpha() else None
    return result
